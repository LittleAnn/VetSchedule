import pandas as pd
import random
import os
import subprocess
import calendar
import openpyxl  # type: ignore
from openpyxl import load_workbook  # type: ignore
from openpyxl.styles import PatternFill  # type: ignore
import time

# --- Define scheduling function ---
def generate_schedule(file_path, save_path, year, month):
    random.seed(time.time())  # Ensure different output every time

    xls = pd.ExcelFile(file_path)
    data = pd.read_excel(xls, sheet_name='Dyspozycje')
    limits = pd.read_excel(xls, sheet_name='Limit zmian')
    preferences = pd.read_excel(xls, sheet_name='Preferencje zmian')
    fixed_df = pd.read_excel(xls, sheet_name='Ustalone zmiany')
    try:
        vacations = pd.read_excel(xls, sheet_name='Urlopy')
        vacation_days = {(row['Pracownik'], int(row['Dzień'])) for idx, row in vacations.iterrows()}
    except:
        vacation_days = set()

    availability_marker = 1

    year = int(year)
    month = int(month)
    cal = calendar.Calendar()
    weekends = [day for day, weekday in cal.itermonthdays2(year, month) if day != 0 and weekday >= 5]

    shift_limits = {}
    for _, row in limits.iterrows():
        employee = row['Pracownik']
        shift_limits[employee] = {
            'day': row['Dzień'],
            'night': row['Noc'],
            'weekend': row['Weekend']
        }

    shift_preferences = {}
    for _, row in preferences.iterrows():
        employee = row['Pracownik']
        shift_preferences[employee] = {
            'day': row['Dzień'],
            'night': row['Noc'],
            'weekend': row['Weekend']
        }

    assigned_shifts = {
        emp: {'day': 0, 'night': 0, 'weekend': 0} for emp in shift_limits
    }
    last_day_shift = {emp: None for emp in shift_limits}
    last_night_shift = {emp: None for emp in shift_limits}

    assignments = []
    max_day = data.shape[0]
    all_employees = list(shift_limits.keys())
    shift_matrix = pd.DataFrame(index=range(1, max_day + 1), columns=all_employees)

    # Initialize with 'Wolne' or 'X' based on availability
    for idx, row in data.iterrows():
        for emp in all_employees:
            shift_matrix.at[idx + 1, emp] = 'Wolne' if row.get(emp, 0) == availability_marker else 'X'

    fixed_assignments = {}
    for _, row in fixed_df.iterrows():
        day = int(row['Dzień'])
        shift_type = row['Typ zmiany']
        emp = row['Pracownik']
        if day not in fixed_assignments:
            fixed_assignments[day] = {}
        fixed_assignments[day][shift_type] = emp
        if emp in assigned_shifts:
            assigned_shifts[emp][shift_type] += 1
            if shift_type == 'night':
                last_night_shift[emp] = day
            else:
                last_day_shift[emp] = day
            shift_matrix.at[day, emp] = shift_type

    for idx, row in data.iterrows():
        day = idx + 1
        is_weekend = day in weekends
        shift_label_display = 'weekend' if is_weekend else 'day'
        shift_type_internal = 'day'  # Always count toward 'day' internally
        shift_type_night = 'night'

        available_employees = [col for col in data.columns[1:] if row[col] == availability_marker]

        # --- Day Shift (includes weekend) ---
        if not (day in fixed_assignments and shift_label_display in fixed_assignments[day]):
            eligible_day = [
                emp for emp in available_employees
                if emp in assigned_shifts
                and assigned_shifts[emp][shift_type_internal] < shift_limits[emp][shift_type_internal]
                and shift_preferences[emp][shift_type_internal] == 1
                and (not is_weekend or shift_preferences[emp]['weekend'] == 1)
                and (last_night_shift[emp] != day - 1)
                and (emp, day) not in vacation_days
            ]

            def day_score(emp):
                return (
                    assigned_shifts[emp][shift_type_internal],
                    sum(assigned_shifts[emp].values()),
                    random.random()
                )

            sorted_day = sorted(eligible_day, key=day_score)
            max_day_workers = 3 if is_weekend else 2
            num_to_assign = min(max_day_workers, len(sorted_day))
            assigned_today = sorted_day[:num_to_assign]

            for emp in assigned_today:
                assigned_shifts[emp][shift_type_internal] += 1
                if is_weekend:
                    assigned_shifts[emp]['weekend'] += 1
                last_day_shift[emp] = day
                shift_matrix.at[day, emp] = shift_label_display

        # --- Night Shift ---
        if not (day in fixed_assignments and shift_type_night in fixed_assignments[day]):
            eligible_night = [
                emp for emp in available_employees
                if emp in assigned_shifts
                and assigned_shifts[emp]['night'] < shift_limits[emp]['night']
                and shift_preferences[emp]['night'] == 1
                and (emp, day) not in vacation_days
            ]

            def night_score(emp):
                return (
                    assigned_shifts[emp]['night'],
                    sum(assigned_shifts[emp].values()),
                    random.random()
                )

            sorted_night = sorted(eligible_night, key=night_score)
            assigned_night = sorted_night[:1]

            for emp in assigned_night:
                assigned_shifts[emp]['night'] += 1
                last_night_shift[emp] = day
                shift_matrix.at[day, emp] = 'night'

    label_translation = {'day': 'Dzień', 'night': 'Noc', 'weekend': 'Dzień'}
    shift_matrix = shift_matrix.applymap(lambda x: label_translation.get(x, x))

    summary_data = []
    for emp, shifts in assigned_shifts.items():
        summary_data.append({
            'Pracownik': emp,
            'Dniowe zmiany': shifts['day'],
            'Nocne zmiany': shifts['night'],
            'Weekendowe zmiany': shifts['weekend'],
            'Suma zmian': sum(shifts.values())
        })
    summary_df = pd.DataFrame(summary_data)

    with pd.ExcelWriter(save_path, engine='openpyxl') as writer:
        shift_matrix.to_excel(writer, sheet_name='Grafik', index_label='Dzień')
        summary_df.to_excel(writer, sheet_name='Podsumowanie', index=False)

    wb = load_workbook(save_path)
    ws = wb['Grafik']

    fill_day = PatternFill(start_color="66B902", end_color="66B902", fill_type="solid")
    fill_night = PatternFill(start_color="2875D7", end_color="2875D7", fill_type="solid")
    fill_free = PatternFill(start_color="E8E8A5", end_color="E8E8A5", fill_type="solid")
    fill_weekend = PatternFill(start_color="C112CD", end_color="C112CD", fill_type="solid")
    fill_unavailable = PatternFill(start_color="BBBBBB", end_color="BBBBBB", fill_type="solid")

    for row in ws.iter_rows(min_row=2, min_col=2):
        day_cell = row[0].row
        is_weekend = (day_cell - 1) in weekends
        for cell in row:
            if cell.value == 'Dzień':
                cell.fill = fill_weekend if is_weekend else fill_day
            elif cell.value == 'Noc':
                cell.fill = fill_night
            elif cell.value == 'Wolne':
                cell.fill = fill_free
            elif cell.value == 'X':
                cell.fill = fill_unavailable

    for column_cells in ws.columns:
        max_length = 0
        column = column_cells[0].column_letter
        for cell in column_cells:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[column].width = max_length + 2

    wb.save(save_path)
